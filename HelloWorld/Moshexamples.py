def emoji_converter(message):
    words = message.split(" ")
    emojis = {
        ":)": "😊",
        ":(": "😒"
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
        return output


message = input(">")
print(emoji_converter(message))

#handling errors
try:
    age = int(input("Age: "))
    print(age)
    income = 20000
    risk = income / age
except ZeroDivisionError:
    print("Age cannot be 0.")
except ValueError:
    print('Invalid value')

#Classes
class Point:
    def move(self):
        print("move")

    def draw(self):
        print("draw")


point1 = Point()
point1.x = 10
point1.y = 20
print(point1.x)
point1.draw()

point2 = Point()
point2.x = 1
print(point2.x)


#Constructors
class Point:
    def __int__(self, x, y):
        self.x = x
        self.y = y

    def move(self):
        print("move")

    def draw(self):
        print("draw")


point = Point()
point.x = 11
print(point.x)


class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f"Hi, I am {self.name}")


john = Person("John Smith")
john.talk()

bob = Person("Bob Smith")
bob.talk()

#Inheritance
class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal):
    def bark(self):
        print("bark")


class Cat(Mammal):
    def be_annoying(self):
        print("annoying")


dog1 = Dog()
dog1.walk()
dog1.bark()
cat1 = Cat()
cat1.be_annoying()

#Modules
from utils import find_max
import utils


numbers = [10, 3, 6, 2]
maximum = find_max(numbers)
print(maximum(numbers))

#Packages
from ecommerce.shipping import calc_shipping

calc_shipping()

#built_in functions
import random

for i in range(3):
    print(random.randint(10, 20))

members = [" john", "mary", "josh"]
leader = random.choice(members)
print(leader)

import random

class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())
#files and directories
from pathlib import Path

path = Path("ecommerce")
print(path.exists())
from pathlib import Path

path = Path("emails")
print(path.mkdir())
from pathlib import Path

path = Path()
for file in (path.glob('*.py')):
    print(file)

for file in (path.glob('*')):
for file in (path.glob('*.xls')):
    import openpyxl as xl

    wb = xl.load_workbook('transactions.xlsx')
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell = sheet.cell(1, 1)
    print(cell.value)
    print(sheet.max_row)



